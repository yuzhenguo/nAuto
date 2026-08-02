import openpyxl
import sys

xlsx_path = r'e:\네이버자동주문\개발문서\결재목록.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
print(f'총 컬럼수: {ws.max_column}, 총 행수: {ws.max_row}')
print('헤더(1행):', [ws.cell(1, c).value for c in range(1, ws.max_column + 1)])
for row_idx in range(2, min(ws.max_row + 1, 5)):
    print(f'{row_idx}행:', [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)])
