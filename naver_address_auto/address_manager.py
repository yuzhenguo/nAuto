"""
address_manager.py
엑셀(주소록.xlsx) 스레드 안전 읽기/쓰기 관리 모듈
컬럼 구조:
  1열: 수취인 이름
  2열: 주소 검색어
  3열: 우편번호
  4열: 휴대폰 번호 (010-xxxx-xxxx)
  5열: 상세 주소
  6열: 작업 상태 (공백=미작업, Y=성공, F=실패)
  7열: 핸드폰 디바이스 ID
  8열: (예비)
  9열: 주소초기화 여부 (Y=기존 배송지 모두 삭제, 공백=유지)
"""

import threading
import openpyxl
import os
from typing import Optional


class AddressRow:
    """주소록 단일 행 데이터"""
    def __init__(self, row_index: int, name: str, address_search: str,
                 zipcode: str, phone: str, detail_address: str,
                 status: str, device_id: str, delete_existing: bool = False):
        self.row_index = row_index            # 엑셀 실제 행 번호 (1-based)
        self.name = name                      # 1열: 수취인
        self.address_search = address_search  # 2열: 주소 검색어
        self.zipcode = str(zipcode).strip() if zipcode else ""   # 3열: 우편번호
        self.phone = str(phone).strip() if phone else ""         # 4열: 휴대폰
        self.detail_address = detail_address  # 5열: 상세주소
        self.status = str(status).strip() if status else ""      # 6열: 상태
        self.device_id = str(device_id).strip() if device_id else ""  # 7열: 기기 ID
        self.delete_existing = delete_existing  # 9열: 기존 주소 삭제 여부

    def get_phone_middle(self) -> str:
        """전화번호 중간 4자리 반환 (앞 3자리 제거)"""
        digits = ''.join(filter(str.isdigit, self.phone))
        if len(digits) >= 11:
            return digits[3:7]   # 01012345678 → 1234
        elif len(digits) >= 8:
            return digits[3:7]
        return ""

    def get_phone_last(self) -> str:
        """전화번호 마지막 4자리 반환"""
        digits = ''.join(filter(str.isdigit, self.phone))
        if len(digits) >= 8:
            return digits[-4:]
        return ""

    def __repr__(self):
        return (f"AddressRow(row={self.row_index}, name={self.name}, "
                f"zipcode={self.zipcode}, device={self.device_id}, status={self.status}, "
                f"delete_existing={self.delete_existing})")


class AddressManager:
    """
    스레드 안전한 엑셀 주소록 관리자
    여러 워커 스레드가 동시에 접근해도 안전하게 처리
    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self._lock = threading.Lock()

    def get_pending_rows_for_device(self, device_id: str) -> list:
        """
        특정 기기 ID에 해당하며 6열이 공백인(미처리) 행 목록 반환
        device_id가 비어있으면 기기 ID 무관하게 공백 행 반환
        """
        with self._lock:
            rows = []
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                for row_idx in range(2, ws.max_row + 1):  # 2행부터 시작
                    name = ws.cell(row_idx, 1).value
                    # 1열(이름)이 비어있으면 데이터 끝
                    if not name or str(name).strip() == "":
                        break

                    status = ws.cell(row_idx, 6).value
                    dev = ws.cell(row_idx, 7).value

                    status_str = str(status).strip() if status else ""
                    dev_str = str(dev).strip() if dev else ""

                    # 6열이 공백이고, 기기 ID가 일치하는 행만
                    if status_str == "" and (device_id == "" or dev_str == device_id):
                        addr_search = ws.cell(row_idx, 2).value
                        zipcode = ws.cell(row_idx, 3).value
                        phone = ws.cell(row_idx, 4).value
                        detail = ws.cell(row_idx, 5).value

                        # 9열 (주소초기화 여부)
                        delete_val = ws.cell(row_idx, 9).value
                        delete_existing = False
                        if delete_val:
                            delete_existing = str(delete_val).strip().upper() == "Y"

                        rows.append(AddressRow(
                            row_index=row_idx,
                            name=str(name).strip(),
                            address_search=str(addr_search).strip() if addr_search else "",
                            zipcode=zipcode,
                            phone=phone,
                            detail_address=str(detail).strip() if detail else "",
                            status=status_str,
                            device_id=dev_str,
                            delete_existing=delete_existing
                        ))
            except Exception as e:
                print(f"[AddressManager] 엑셀 읽기 오류: {e}")
            return rows

    def get_next_pending_row(self, device_id: str) -> Optional[AddressRow]:
        """기기 ID에 해당하는 미처리 행 중 첫 번째 반환"""
        rows = self.get_pending_rows_for_device(device_id)
        return rows[0] if rows else None

    def mark_success(self, row_index: int):
        """해당 행의 6열을 Y로 업데이트"""
        self._update_status(row_index, "Y")

    def mark_failed(self, row_index: int):
        """해당 행의 6열을 F로 업데이트"""
        self._update_status(row_index, "F")

    def _update_status(self, row_index: int, status: str):
        """6열 상태값 업데이트 (스레드 안전)"""
        with self._lock:
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                ws.cell(row_index, 6).value = status
                wb.save(self.xlsx_path)
            except Exception as e:
                print(f"[AddressManager] 엑셀 쓰기 오류 (row={row_index}): {e}")

    def has_pending_rows(self, device_id: str) -> bool:
        """미처리 행이 남아있는지 확인"""
        return len(self.get_pending_rows_for_device(device_id)) > 0

    def get_all_rows_summary(self) -> dict:
        """전체 현황 요약 반환 (GUI 표시용)"""
        with self._lock:
            summary = {"total": 0, "done": 0, "failed": 0, "pending": 0}
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                for row_idx in range(2, ws.max_row + 1):
                    name = ws.cell(row_idx, 1).value
                    if not name or str(name).strip() == "":
                        break
                    summary["total"] += 1
                    status = ws.cell(row_idx, 6).value
                    status_str = str(status).strip() if status else ""
                    if status_str == "Y":
                        summary["done"] += 1
                    elif status_str == "F":
                        summary["failed"] += 1
                    else:
                        summary["pending"] += 1
            except Exception as e:
                print(f"[AddressManager] 현황 조회 오류: {e}")
            return summary
