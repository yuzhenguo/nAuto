"""
order_manager.py
결재목록.xlsx 스레드 안전 읽기/쓰기 관리 모듈

컬럼 구조 (헤더로 자동 감지):
  검색어      : 상품 검색 키워드
  판매자명    : 스토어/판매자 이름 (매칭용)
  상품명      : 상품명 (포함 여부로 매칭)
  수취인      : 배송지 수취인명
  전화번호    : 수취인 전화번호
  비밀번호    : 결제 비밀번호 (숫자, 예: 123456)
  2차비밀번호 : 현대카드 2차 비밀번호 (4자리)
  완료여부    : 공백=미처리, Y=완료, F=실패

헤더가 없거나 컬럼명이 다를 경우 컬럼 인덱스로 직접 지정 가능 (아래 COL_* 상수 참고)
"""

import threading
import openpyxl
import os
from typing import Optional, List


# ─── 컬럼 인덱스 기본값 (헤더 자동 감지 실패 시 사용, 1-based) ───────────────
# 헤더 자동 감지 우선, 없으면 아래 기본값 사용
COL_SEARCH_KEYWORD  = 1   # 검색어
COL_SELLER_NAME     = 2   # 판매자명
COL_PRODUCT_NAME    = 3   # 상품명
COL_RECIPIENT_NAME  = 4   # 수취인
COL_PHONE           = 5   # 전화번호
COL_PASSWORD        = 6   # 비밀번호
COL_STATUS          = 7   # 완료여부
COL_PAYMENT_METHOD  = 8   # 결제방식
COL_DEVICE_ID       = 9   # 폰ID (기기 ID)
COL_LOGIN_ID        = 11  # 로그인아이디
COL_SECOND_PASSWORD = 12  # 2차비밀번호 (현대카드)


# ─── 헤더 키워드 매핑 (대소문자/공백 무시, 긴 키워드 우선) ───────────────────
HEADER_KEYWORDS = {
    "search_keyword": ["검색어", "searchkeyword", "search_keyword", "keyword"],
    "seller_name":    ["판매자명", "판매자", "seller", "스토어명"],
    "product_name":   ["매칭상품명", "매칭 상품명", "상품명", "productname"],
    "recipient_name": ["수취인", "수령인", "recipient", "받는사람", "받는분"],
    "phone":          ["전화번호", "연락처", "phone"],
    "password":       ["비밀번호", "password", "암호"],
    "status":         ["완료여부", "처리여부", "작업여부", "status"],
    "payment_method": ["결제방식", "결재방식", "payment"],
    "device_id":      ["폰id", "기기id", "단말기id", "deviceid", "device_id"],
    "login_id":       ["로그인아이디", "로그인id", "loginid", "login_id"],
    "second_password": ["2차비밀번호", "2차암호", "2차비번", "secondarypassword"],
}


def _norm_device_id(val: str) -> str:
    return str(val or "").strip().upper()


def _device_ids_match(excel_id: str, device_id: str) -> bool:
    """엑셀 폰ID와 ADB 기기ID 비교 (대소문자/공백 무시). 엑셀이 비어 있으면 미지정으로 간주."""
    a, b = _norm_device_id(excel_id), _norm_device_id(device_id)
    if not b:
        return True
    if not a:
        return True  # 폰ID 공백 행은 현재 기기가 가져갈 수 있음
    return a == b


class OrderRow:
    """결재목록 단일 행 데이터"""

    def __init__(self, row_index: int,
                 search_keyword: str,
                 seller_name: str,
                 product_name: str,
                 recipient_name: str,
                 phone: str,
                 password: str,
                 status: str,
                 payment_method: str = "",
                 device_id: str = "",
                 login_id: str = "",
                 second_password: str = ""):
        self.row_index      = row_index         # 엑셀 실제 행 번호 (1-based)
        self.search_keyword = search_keyword    # 검색어
        self.seller_name    = seller_name       # 판매자명
        self.product_name   = product_name      # 상품명
        self.recipient_name = recipient_name    # 수취인
        self.phone          = str(phone).strip() if phone else ""   # 전화번호
        self.password       = str(password).strip() if password else ""  # 비밀번호
        self.status         = str(status).strip() if status else ""  # 완료여부
        self.payment_method = str(payment_method).strip() if payment_method else "" # 결제방식
        self.device_id      = str(device_id).strip() if device_id else ""  # 폰ID
        self.login_id       = str(login_id).strip() if login_id else "" # 로그인아이디
        self.second_password = str(second_password).strip() if second_password else ""  # 2차비밀번호

    def get_phone_digits(self) -> str:
        """전화번호에서 숫자만 추출"""
        return ''.join(filter(str.isdigit, self.phone))

    def get_password_digits(self) -> str:
        """비밀번호 숫자 문자열"""
        return ''.join(filter(str.isdigit, self.password))

    def get_second_password_digits(self) -> str:
        """현대카드 2차비밀번호 숫자 문자열"""
        return ''.join(filter(str.isdigit, self.second_password))

    def __repr__(self):
        return (f"OrderRow(row={self.row_index}, keyword={self.search_keyword!r}, "
                f"seller={self.seller_name!r}, product={self.product_name!r}, "
                f"recipient={self.recipient_name!r}, status={self.status!r}, "
                f"device_id={self.device_id!r}, login_id={self.login_id!r})")


def _detect_columns(ws) -> dict:
    """
    1행을 헤더로 읽어 컬럼 인덱스 자동 감지.
    짧은 부분문자열 오탐 방지를 위해 정확 일치 > 긴 키워드 포함 순으로 채점.
    감지 실패 컬럼은 COL_* 기본값 사용.
    """
    mapping = {
        "search_keyword": COL_SEARCH_KEYWORD,
        "seller_name":    COL_SELLER_NAME,
        "product_name":   COL_PRODUCT_NAME,
        "recipient_name": COL_RECIPIENT_NAME,
        "phone":          COL_PHONE,
        "password":       COL_PASSWORD,
        "status":         COL_STATUS,
        "payment_method": COL_PAYMENT_METHOD,
        "device_id":      COL_DEVICE_ID,
        "login_id":       COL_LOGIN_ID,
        "second_password": COL_SECOND_PASSWORD,
    }
    scores = {k: -1 for k in mapping}

    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(1, col_idx).value
        if not cell_val:
            continue
        cell_str = str(cell_val).strip().lower().replace(" ", "").replace("_", "")
        for field, keywords in HEADER_KEYWORDS.items():
            # '2차비밀번호'가 일반 '비밀번호'로 오탐되지 않도록
            if field == "password" and "2차" in cell_str:
                continue
            for kw in keywords:
                kw_n = kw.lower().replace(" ", "").replace("_", "")
                if not kw_n:
                    continue
                if cell_str == kw_n:
                    score = 1000 + len(kw_n)
                elif kw_n in cell_str:
                    score = len(kw_n)
                else:
                    continue
                if score > scores[field]:
                    scores[field] = score
                    mapping[field] = col_idx

    return mapping


class OrderManager:
    """
    스레드 안전한 결재목록 엑셀 관리자.
    헤더 자동 감지 + COL_* 상수 폴백.
    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self._lock = threading.Lock()
        self._col_map: Optional[dict] = None  # 캐시

    def _get_col_map(self, ws) -> dict:
        if self._col_map is None:
            self._col_map = _detect_columns(ws)
        return self._col_map

    def _str(self, val) -> str:
        return str(val).strip() if val is not None else ""

    def _str_pin(self, val) -> str:
        """엑셀 숫자셀(1234.0)이 비밀번호로 깨지지 않게 변환."""
        if val is None:
            return ""
        if isinstance(val, bool):
            return ""
        if isinstance(val, int):
            return str(val)
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            return str(val).strip()
        return str(val).strip()

    def get_pending_rows(self, device_id: str = "") -> List[OrderRow]:
        """완료여부가 공백인 미처리 행 목록 반환.
        device_id가 지정되면 엑셀의 폰ID 컬럼 값과 일치하는 행만 반환.
        """
        with self._lock:
            rows = []
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                cm = self._get_col_map(ws)

                # 헤더 여부 확인 (1행이 헤더이면 2행부터 시작)
                start_row = 2 if ws.max_row > 1 else 1
                # 1행 첫 셀이 숫자면 헤더 없음 → 1행부터 처리
                first_cell = ws.cell(1, cm["search_keyword"]).value
                if first_cell and str(first_cell).strip().isdigit():
                    start_row = 1

                for row_idx in range(start_row, ws.max_row + 1):
                    # 검색어 컬럼이 비면 데이터 끝
                    keyword_val = ws.cell(row_idx, cm["search_keyword"]).value
                    if not keyword_val or self._str(keyword_val) == "":
                        break

                    status_val = ws.cell(row_idx, cm["status"]).value
                    status_str = self._str(status_val)

                    # 완료여부가 공백/None인 행만 처리 (Y, F, 기타 값은 모두 건너뜀)
                    if status_val is None or status_str.upper() in ("NONE", ""):
                        pass  # 미처리 행 → 작업 대상
                    else:
                        continue  # 이미 처리된 행(Y/F/기타) → 건너뜀

                    # 폰ID 필터링: 대소문자 무시, 엑셀 폰ID가 비어 있으면 현재 기기에 할당 가능
                    row_device_id = ""
                    if "device_id" in cm:
                        row_device_id = self._str(ws.cell(row_idx, cm["device_id"]).value)
                    if device_id and not _device_ids_match(row_device_id, device_id):
                        continue

                    rows.append(OrderRow(
                        row_index      = row_idx,
                        search_keyword = self._str(ws.cell(row_idx, cm["search_keyword"]).value),
                        seller_name    = self._str(ws.cell(row_idx, cm["seller_name"]).value),
                        product_name   = self._str(ws.cell(row_idx, cm["product_name"]).value),
                        recipient_name = self._str(ws.cell(row_idx, cm["recipient_name"]).value),
                        phone          = self._str(ws.cell(row_idx, cm["phone"]).value),
                        password       = self._str_pin(ws.cell(row_idx, cm["password"]).value),
                        status         = status_str,
                        payment_method = self._str(ws.cell(row_idx, cm["payment_method"]).value),
                        device_id      = row_device_id,
                        login_id       = self._str(ws.cell(row_idx, cm["login_id"]).value),
                        second_password = self._str_pin(ws.cell(row_idx, cm.get("second_password", COL_SECOND_PASSWORD)).value),
                    ))
            except Exception as e:
                print(f"[OrderManager] 엑셀 읽기 오류: {e}")
            return rows

    def get_next_pending(self, device_id: str = "") -> Optional[OrderRow]:
        """미처리 행 중 첫 번째 반환. device_id를 지정하면 해당 폰ID 행만 대상."""
        rows = self.get_pending_rows(device_id=device_id)
        return rows[0] if rows else None

    def describe_pending_filter(self, device_id: str) -> str:
        """기기 필터로 0건일 때 원인 파악용 요약 문자열"""
        all_pending = self.get_pending_rows()
        matched = self.get_pending_rows(device_id=device_id)
        excel_ids = sorted({(r.device_id or "(빈값)") for r in all_pending})
        return (
            f"엑셀 pending={len(all_pending)}건, "
            f"기기 '{device_id}' 매칭={len(matched)}건, "
            f"엑셀 폰ID 목록={excel_ids}"
        )

    def mark_success(self, row_index: int):
        """해당 행 완료여부를 Y로 업데이트"""
        self._update_status(row_index, "Y")

    def mark_failed(self, row_index: int):
        """해당 행 완료여부를 F로 업데이트"""
        self._update_status(row_index, "F")

    def _update_status(self, row_index: int, status: str):
        """완료여부 컬럼 업데이트 (스레드 안전)"""
        with self._lock:
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                cm = self._get_col_map(ws)
                ws.cell(row_index, cm["status"]).value = status
                wb.save(self.xlsx_path)
            except Exception as e:
                print(f"[OrderManager] 엑셀 쓰기 오류 (row={row_index}): {e}")

    def get_summary(self) -> dict:
        """전체 현황 요약"""
        with self._lock:
            summary = {"total": 0, "done": 0, "failed": 0, "pending": 0}
            try:
                wb = openpyxl.load_workbook(self.xlsx_path)
                ws = wb.active
                cm = self._get_col_map(ws)
                start_row = 2
                first_cell = ws.cell(1, cm["search_keyword"]).value
                if first_cell and str(first_cell).strip().isdigit():
                    start_row = 1
                for row_idx in range(start_row, ws.max_row + 1):
                    kw = ws.cell(row_idx, cm["search_keyword"]).value
                    if not kw or str(kw).strip() == "":
                        break
                    summary["total"] += 1
                    st = str(ws.cell(row_idx, cm["status"]).value or "").strip()
                    if st == "Y":
                        summary["done"] += 1
                    elif st == "F":
                        summary["failed"] += 1
                    else:
                        summary["pending"] += 1
            except Exception as e:
                print(f"[OrderManager] 현황 조회 오류: {e}")
            return summary
